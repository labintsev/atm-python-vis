import datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class ExcelTemplateGenerator:
    """Create an Excel schedule from a template and a schedule DataFrame."""

    TEMPLATE_COL_START = 7  # G
    TEMPLATE_COL_END = 38
    TEMPLATE_ROW_START = 5
    TEMPLATE_ROW_END = 52

    def __init__(
        self,
        df: pd.DataFrame,
        template_path: Optional[str] = None,
        output_path: Optional[str] = None,
    ):
        if df is None or df.empty:
            raise ValueError("DataFrame не может быть пустым")

        self.df = df.copy()
        self.template_path = template_path 
        self.output_path = output_path 
        self.wb = None
        self.ws = None
        self.date_to_col = {}
        self.time_to_row = {}

    @staticmethod
    def get_fill(blockload: float) -> PatternFill:
        if pd.isna(blockload):
            color = "FFFFFF"
        elif blockload < 0.10:
            color = "E2F0D9"  # light green
        elif blockload < 0.20:
            color = "A9D18E"  # green
        elif blockload < 0.35:
            color = "FFD966"  # yellow
        elif blockload < 0.50:
            color = "F4B183"  # orange
        else:
            color = "FF6666"  # red

        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    def load_template(self) -> "ExcelTemplateGenerator":
        self.wb = load_workbook(self.template_path)
        self.ws = self.wb.active
        return self

    def normalize_schedule(self) -> "ExcelTemplateGenerator":
        if "BlockLoad" not in self.df.columns:
            self.df["BlockLen"] = self.df["Stop"] - self.df["Start"]
            self.df = (
                self.df.groupby(["SchDate", "Start"], as_index=False)
                .agg(
                    BlockLen=("BlockLen", "first"),
                    TotalRealDur=("RealDur", "sum"),
                )
            )
            self.df["BlockLoad"] = self.df["TotalRealDur"] / self.df["BlockLen"]
        return self

    def prepare_dates(self) -> "ExcelTemplateGenerator":
        dates = sorted(pd.to_datetime(self.df["SchDate"]).dt.date.unique())
        self.date_to_col = {}
        for idx, date_value in enumerate(dates):
            col = self.TEMPLATE_COL_START + idx
            self.ws.cell(row=4, column=col).value = str(date_value)
            self.date_to_col[date_value] = col
        return self

    def prepare_time_labels(self) -> "ExcelTemplateGenerator":
        self.time_to_row = {}
        row = self.TEMPLATE_ROW_START
        for hour in range(24):
            for minute in (0, 30):
                t0 = datetime.time(hour=hour, minute=minute)
                t1 = (
                    datetime.datetime.combine(datetime.date.today(), t0)
                    + datetime.timedelta(minutes=30)
                ).time()
                time_str = f"{t0.strftime('%H:%M')} - {t1.strftime('%H:%M')}"
                self.ws.cell(row=row, column=1).value = time_str
                self.time_to_row[(hour, minute)] = row
                row += 1
        return self

    def clear_range(self) -> "ExcelTemplateGenerator":
        fill_none = PatternFill(fill_type=None)
        for row in range(self.TEMPLATE_ROW_START, self.TEMPLATE_ROW_END + 1):
            for col in range(self.TEMPLATE_COL_START, self.TEMPLATE_COL_END + 1):
                self.ws.cell(row=row, column=col).value = None
                self.ws.cell(row=row, column=col).fill = fill_none
        return self

    def fill_colors(self) -> "ExcelTemplateGenerator":
        for _, row in self.df.iterrows():
            date_key = pd.to_datetime(row["SchDate"]).date()
            if date_key not in self.date_to_col:
                continue

            col = self.date_to_col[date_key]
            start_seconds = int(row["Start"])
            total_minutes = start_seconds // 60
            hour = total_minutes // 60
            minute = total_minutes % 60
            minute = 30 if minute >= 30 else 0

            if (hour, minute) not in self.time_to_row:
                continue

            cell_row = self.time_to_row[(hour, minute)]
            blockload = float(row["BlockLoad"])
            self.ws.cell(row=cell_row, column=col).fill = self.get_fill(blockload)
        return self

    def save(self) -> Path:
        output_path = Path(self.output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self.wb.save(output_path)
        return output_path

    def generate(self) -> Path:
        self.load_template()
        self.normalize_schedule()
        self.prepare_dates()
        self.prepare_time_labels()
        self.clear_range()
        self.fill_colors()
        return self.save()


if __name__ == "__main__":
    df = pd.read_csv("data/scheds_orders_prepared.csv")
    generator = ExcelTemplateGenerator(df, output_path="data/МП-.xlsx")
    saved_path = generator.generate()
    print(f"Saved: {saved_path}")
